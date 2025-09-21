from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


HEADERS = [
    "Город", "Дата", "Время", "Температура",
    "Давление", "Влажность", "Погодные явления",
    "Средняя температура", "Магнитное поле", "Примечание"
]


class ExcelWriterError(Exception):
    """Базовое исключение для ошибок создания Excel файла"""
    pass


class ExcelWriter:

    def __init__(self, weather_data: dict, city: str):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.weather_data = weather_data
        self.city = city
        self.ws.title = f"Прогноз погоды в городе {self.city}"

    def _set_headers(self):
        """Заполняем заголовки таблицы"""
        for col, h in enumerate(HEADERS, start=1):
            cell = self.ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _fill_data(self):
        """Заполняем таблицу данными"""
        row = 2
        day_periods = ["Утро", "День", "Вечер", "Ночь"]

        try:

            for date, info in self.weather_data.items():
                periods = info["periods_info"]
                whole = info["whole_day_info"]

                start_row = row
                for period in day_periods:
                    pdata = periods.get(period)
                    if not pdata:
                        continue
                    self.ws.cell(row=row, column=1, value=self.city)
                    self.ws.cell(row=row, column=2, value=date)
                    self.ws.cell(row=row, column=3, value=period)
                    self.ws.cell(row=row, column=4, value=pdata.get("temp"))
                    self.ws.cell(row=row, column=5, value=pdata.get("pressure"))
                    self.ws.cell(row=row, column=6, value=pdata.get("humidity"))
                    self.ws.cell(row=row, column=7, value=pdata.get("text"))
                    row += 1

                for col, key in zip([8, 9, 10], ["avg_day_temp", "mag_field", "warning"]):
                    value = whole.get(key)
                    if value is not None:
                        self.ws.merge_cells(
                            start_row=start_row, start_column=col,
                            end_row=start_row + 3, end_column=col
                        )
                        cell = self.ws.cell(row=start_row, column=col, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        except Exception as e:
            raise ExcelWriterError(f"Критическая ошибка при заполнении данных: {str(e)}")

    def _set_style(self):
        """Задаем ширину колонок и выравниваем текст для улучшения читаемости итогового документа"""
        widths = [18, 14, 9, 16, 14, 14, 27, 27, 18, 23]
        for i, w in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=1, max_col=len(widths)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def process_excel_file(self, filename: str = None):
        """Формируем Excel-файл с отчетом о погоде"""
        try:
            self._set_headers()
            self._fill_data()
            self._set_style()
            if filename:
                self.wb.save(filename)
            return self.wb
        except Exception as e:
            raise ExcelWriterError(f"Ошибка при создании Excel файла: {str(e)}")
